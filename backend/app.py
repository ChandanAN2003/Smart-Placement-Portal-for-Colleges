"""
Main Flask application for College Placement Management Portal
"""
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
from pathlib import Path
from datetime import datetime, date
import json
from functools import wraps
import PyPDF2
from docx import Document
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO

import sys
from pathlib import Path

# Add backend directory to path
backend_dir = Path(__file__).parent
sys.path.insert(0, str(backend_dir))

from database import db
from gemini_ai import analyze_resume, generate_email_content
from mail_utils import init_mail, send_application_update_email

app = Flask(__name__, template_folder='../frontend/templates', static_folder='../frontend/static')
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')

# Configuration
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'frontend', 'static', 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'doc'}
MAX_UPLOAD_SIZE = int(os.getenv('MAX_UPLOAD_SIZE', 16777216))  # 16MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'resumes'), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'offers'), exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Initialize mail
init_mail(app)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_file(file_path):
    """Extract text from PDF or DOCX file"""
    try:
        if file_path.endswith('.pdf'):
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                text = ''
                for page in pdf_reader.pages:
                    text += page.extract_text()
                return text
        elif file_path.endswith('.docx') or file_path.endswith('.doc'):
            doc = Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            return text
    except Exception as e:
        print(f"Error extracting text: {e}")
        return ""
    return ""

def login_required(f):
    """Decorator for routes that require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(*roles):
    """Decorator for routes that require specific role(s)"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('index'))
            if session.get('role') not in roles:
                flash('You do not have permission to access this page.', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ==================== Routes ====================

@app.route('/')
def index():
    """Home page"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration"""
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        department = request.form.get('department', '')
        
        if not all([name, email, password, role]):
            flash('All fields are required.', 'error')
            return render_template('index.html', show_register=True)
        
        # Check if email exists
        existing = db.execute_query(
            "SELECT id FROM users WHERE email = %s",
            (email,),
            fetch_one=True
        )
        
        if existing:
            flash('Email already registered.', 'error')
            return render_template('index.html', show_register=True)
        
        # Create user
        password_hash = generate_password_hash(password)
        is_approved = True if role == 'tpo' else False
        
        db.execute_query(
            "INSERT INTO users (name, email, password_hash, role, department, is_approved) VALUES (%s, %s, %s, %s, %s, %s)",
            (name, email, password_hash, role, department, is_approved)
        )
        
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('index'))
    
    return render_template('index.html', show_register=True)

@app.route('/login', methods=['POST'])
def login():
    """User login"""
    email = request.form.get('email')
    password = request.form.get('password')
    
    if not email or not password:
        flash('Email and password are required.', 'error')
        return redirect(url_for('index'))
    
    user = db.execute_query(
        "SELECT * FROM users WHERE email = %s",
        (email,),
        fetch_one=True
    )
    
    if user and check_password_hash(user['password_hash'], password):
        if not user['is_approved'] and user['role'] != 'tpo':
            flash('Your account is pending approval from HOD.', 'warning')
            return redirect(url_for('index'))
        
        session['user_id'] = user['id']
        session['name'] = user['name']
        session['email'] = user['email']
        session['role'] = user['role']
        session['department'] = user.get('department', '')
        
        flash(f'Welcome, {user["name"]}!', 'success')
        return redirect(url_for('dashboard'))
    else:
        flash('Invalid email or password.', 'error')
        return redirect(url_for('index'))

@app.route('/logout')
def logout():
    """User logout"""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Role-based dashboard redirect"""
    role = session.get('role')
    
    if role == 'student':
        return redirect(url_for('student_dashboard'))
    elif role == 'hod':
        return redirect(url_for('hod_dashboard'))
    elif role == 'tpo':
        return redirect(url_for('tpo_dashboard'))
    
    return redirect(url_for('index'))

# ==================== Student Routes ====================

@app.route('/student/dashboard')
@login_required
@role_required('student')
def student_dashboard():
    """Student dashboard"""
    user_id = session['user_id']
    
    # Get student info
    student = db.execute_query(
        "SELECT * FROM users WHERE id = %s",
        (user_id,),
        fetch_one=True
    )
    
    # Get active drives
    drives = db.execute_query(
        "SELECT * FROM drives WHERE status = 'active' AND last_date >= CURDATE() ORDER BY last_date ASC",
        fetch_all=True
    )
    
    # Get applications
    applications = db.execute_query(
        """SELECT a.*, d.company_name, d.job_role, d.last_date 
           FROM applications a 
           JOIN drives d ON a.drive_id = d.id 
           WHERE a.student_id = %s 
           ORDER BY a.applied_at DESC""",
        (user_id,),
        fetch_all=True
    )
    
    # Get resume
    resume = db.execute_query(
        "SELECT * FROM resumes WHERE user_id = %s ORDER BY analyzed_at DESC LIMIT 1",
        (user_id,),
        fetch_one=True
    )
    
    # Get notifications
    notifications = db.execute_query(
        "SELECT * FROM notifications WHERE user_id = %s AND is_read = FALSE ORDER BY created_at DESC LIMIT 10",
        (user_id,),
        fetch_all=True
    )
    
    return render_template('student_dashboard.html',
                         student=student,
                         drives=drives or [],
                         applications=applications or [],
                         resume=resume,
                         notifications=notifications or [])

@app.route('/student/upload_resume', methods=['POST'])
@login_required
@role_required('student')
def upload_resume():
    """Upload and analyze resume"""
    if 'resume' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('student_dashboard'))
    
    file = request.files['resume']
    if file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('student_dashboard'))
    
    if not allowed_file(file.filename):
        flash('Invalid file type. Please upload PDF or DOCX.', 'error')
        return redirect(url_for('student_dashboard'))
    
    # Check file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > MAX_UPLOAD_SIZE:
        flash('File size exceeds maximum limit (16MB).', 'error')
        return redirect(url_for('student_dashboard'))
    
    # Save file
    filename = secure_filename(f"{session['user_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'resumes', filename)
    file.save(file_path)
    
    # Extract text
    resume_text = extract_text_from_file(file_path)
    
    if not resume_text:
        flash('Could not extract text from resume. Please ensure the file is not corrupted.', 'error')
        return redirect(url_for('student_dashboard'))
    
    # Analyze with Gemini (basic analysis for now)
    analysis = analyze_resume(resume_text, "General")
    
    # Save to database
    job_fit_score = analysis.get('job_fit_score', 0)
    feedback = json.dumps(analysis)
    
    db.execute_query(
        "INSERT INTO resumes (user_id, file_path, original_filename, job_fit_score, feedback) VALUES (%s, %s, %s, %s, %s)",
        (session['user_id'], file_path, file.filename, job_fit_score, feedback)
    )
    
    flash('Resume uploaded and analyzed successfully!', 'success')
    return redirect(url_for('student_dashboard'))

@app.route('/student/apply/<int:drive_id>', methods=['POST'])
@login_required
@role_required('student')
def apply_drive(drive_id):
    """Apply for a placement drive"""
    user_id = session['user_id']
    
    # Check if already applied
    existing = db.execute_query(
        "SELECT id FROM applications WHERE student_id = %s AND drive_id = %s",
        (user_id, drive_id),
        fetch_one=True
    )
    
    if existing:
        flash('You have already applied for this drive.', 'warning')
        return redirect(url_for('student_dashboard'))
    
    # Check if resume exists
    resume = db.execute_query(
        "SELECT id FROM resumes WHERE user_id = %s",
        (user_id,),
        fetch_one=True
    )
    
    if not resume:
        flash('Please upload your resume before applying.', 'error')
        return redirect(url_for('student_dashboard'))
    
    # Create application
    db.execute_query(
        "INSERT INTO applications (student_id, drive_id, status) VALUES (%s, %s, %s)",
        (user_id, drive_id, 'Applied')
    )
    
    # Create notification
    drive = db.execute_query(
        "SELECT company_name, job_role FROM drives WHERE id = %s",
        (drive_id,),
        fetch_one=True
    )
    
    db.execute_query(
        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
        (user_id, f"Application submitted for {drive['company_name']} - {drive['job_role']}", 'success')
    )
    
    flash('Application submitted successfully!', 'success')
    return redirect(url_for('student_dashboard'))

@app.route('/student/resume_analysis/<int:drive_id>')
@login_required
@role_required('student')
def resume_analysis(drive_id):
    """Get AI resume analysis for a specific drive"""
    user_id = session['user_id']
    
    # Get resume
    resume = db.execute_query(
        "SELECT * FROM resumes WHERE user_id = %s ORDER BY analyzed_at DESC LIMIT 1",
        (user_id,),
        fetch_one=True
    )
    
    if not resume:
        return jsonify({'error': 'No resume found'}), 404
    
    # Get drive details
    drive = db.execute_query(
        "SELECT * FROM drives WHERE id = %s",
        (drive_id,),
        fetch_one=True
    )
    
    if not drive:
        return jsonify({'error': 'Drive not found'}), 404
    
    # Extract resume text
    resume_text = extract_text_from_file(resume['file_path'])
    
    # Analyze with Gemini for this specific job
    analysis = analyze_resume(resume_text, drive['job_role'], drive.get('job_description', ''))
    
    return jsonify(analysis)

# ==================== HOD Routes ====================

@app.route('/hod/dashboard')
@login_required
@role_required('hod')
def hod_dashboard():
    """HOD dashboard"""
    department = session.get('department', '')
    
    # Get pending approvals
    pending_students = db.execute_query(
        "SELECT * FROM users WHERE role = 'student' AND department = %s AND is_approved = FALSE",
        (department,),
        fetch_all=True
    )
    
    # Get department statistics
    total_students = db.execute_query(
        "SELECT COUNT(*) as count FROM users WHERE role = 'student' AND department = %s",
        (department,),
        fetch_one=True
    )
    
    approved_students = db.execute_query(
        "SELECT COUNT(*) as count FROM users WHERE role = 'student' AND department = %s AND is_approved = TRUE",
        (department,),
        fetch_one=True
    )
    
    # Get department applications
    applications = db.execute_query(
        """SELECT a.*, u.name as student_name, u.email, d.company_name, d.job_role 
           FROM applications a 
           JOIN users u ON a.student_id = u.id 
           JOIN drives d ON a.drive_id = d.id 
           WHERE u.department = %s 
           ORDER BY a.applied_at DESC 
           LIMIT 20""",
        (department,),
        fetch_all=True
    )
    
    return render_template('hod_dashboard.html',
                         pending_students=pending_students or [],
                         total_students=total_students['count'] if total_students else 0,
                         approved_students=approved_students['count'] if approved_students else 0,
                         applications=applications or [])

@app.route('/hod/approve_student/<int:student_id>', methods=['POST'])
@login_required
@role_required('hod')
def approve_student(student_id):
    """Approve a student"""
    db.execute_query(
        "UPDATE users SET is_approved = TRUE WHERE id = %s AND role = 'student'",
        (student_id,)
    )
    
    # Create notification
    student = db.execute_query(
        "SELECT name FROM users WHERE id = %s",
        (student_id,),
        fetch_one=True
    )
    
    db.execute_query(
        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
        (student_id, "Your account has been approved by HOD. You can now access all features.", 'success')
    )
    
    flash('Student approved successfully.', 'success')
    return redirect(url_for('hod_dashboard'))

@app.route('/hod/reject_student/<int:student_id>', methods=['POST'])
@login_required
@role_required('hod')
def reject_student(student_id):
    """Reject a student (delete account)"""
    db.execute_query(
        "DELETE FROM users WHERE id = %s AND role = 'student'",
        (student_id,)
    )
    
    flash('Student account removed.', 'info')
    return redirect(url_for('hod_dashboard'))

@app.route('/hod/export_report')
@login_required
@role_required('hod')
def export_hod_report():
    """Export department report as Excel"""
    department = session.get('department', '')
    
    # Get all students and their applications
    students = db.execute_query(
        """SELECT u.*, 
                  COUNT(DISTINCT a.id) as total_applications,
                  COUNT(DISTINCT CASE WHEN a.status = 'Selected' THEN a.id END) as selected_count
           FROM users u
           LEFT JOIN applications a ON u.id = a.student_id
           WHERE u.role = 'student' AND u.department = %s
           GROUP BY u.id""",
        (department,),
        fetch_all=True
    )
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Department Report"
    
    # Headers
    ws.append(['Name', 'Email', 'Approved', 'Total Applications', 'Selected'])
    
    # Data
    for student in students:
        ws.append([
            student['name'],
            student['email'],
            'Yes' if student['is_approved'] else 'No',
            student['total_applications'] or 0,
            student['selected_count'] or 0
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'{department}_report.xlsx')

# ==================== TPO Routes ====================

@app.route('/tpo/dashboard')
@login_required
@role_required('tpo')
def tpo_dashboard():
    """TPO/Admin dashboard"""
    # Get statistics
    total_students = db.execute_query(
        "SELECT COUNT(*) as count FROM users WHERE role = 'student'",
        fetch_one=True
    )
    
    total_drives = db.execute_query(
        "SELECT COUNT(*) as count FROM drives",
        fetch_one=True
    )
    
    active_drives = db.execute_query(
        "SELECT COUNT(*) as count FROM drives WHERE status = 'active' AND last_date >= CURDATE()",
        fetch_one=True
    )
    
    total_applications = db.execute_query(
        "SELECT COUNT(*) as count FROM applications",
        fetch_one=True
    )
    
    selected_count = db.execute_query(
        "SELECT COUNT(*) as count FROM applications WHERE status = 'Selected'",
        fetch_one=True
    )
    
    # Get recent drives
    drives = db.execute_query(
        "SELECT * FROM drives ORDER BY created_at DESC LIMIT 10",
        fetch_all=True
    )
    
    # Get recent applications
    applications = db.execute_query(
        """SELECT a.*, u.name as student_name, u.email, d.company_name, d.job_role 
           FROM applications a 
           JOIN users u ON a.student_id = u.id 
           JOIN drives d ON a.drive_id = d.id 
           ORDER BY a.applied_at DESC 
           LIMIT 20""",
        fetch_all=True
    )
    
    return render_template('tpo_dashboard.html',
                         total_students=total_students['count'] if total_students else 0,
                         total_drives=total_drives['count'] if total_drives else 0,
                         active_drives=active_drives['count'] if active_drives else 0,
                         total_applications=total_applications['count'] if total_applications else 0,
                         selected_count=selected_count['count'] if selected_count else 0,
                         drives=drives or [],
                         applications=applications or [])

@app.route('/tpo/create_drive', methods=['POST'])
@login_required
@role_required('tpo')
def create_drive():
    """Create a new placement drive"""
    company_name = request.form.get('company_name')
    job_role = request.form.get('job_role')
    job_description = request.form.get('job_description', '')
    eligibility = request.form.get('eligibility', '')
    last_date = request.form.get('last_date')
    
    if not all([company_name, job_role, last_date]):
        flash('Company name, job role, and last date are required.', 'error')
        return redirect(url_for('tpo_dashboard'))
    
    db.execute_query(
        "INSERT INTO drives (company_name, job_role, job_description, eligibility, last_date, created_by) VALUES (%s, %s, %s, %s, %s, %s)",
        (company_name, job_role, job_description, eligibility, last_date, session['user_id'])
    )
    
    flash('Placement drive created successfully!', 'success')
    return redirect(url_for('tpo_dashboard'))

@app.route('/tpo/update_application_status/<int:app_id>', methods=['POST'])
@login_required
@role_required('tpo')
def update_application_status(app_id):
    """Update application status"""
    status = request.form.get('status')
    
    if status not in ['Applied', 'Shortlisted', 'Selected', 'Rejected']:
        flash('Invalid status.', 'error')
        return redirect(url_for('tpo_dashboard'))
    
    # Get application details
    application = db.execute_query(
        """SELECT a.*, u.email, u.name as student_name, d.company_name, d.job_role 
           FROM applications a 
           JOIN users u ON a.student_id = u.id 
           JOIN drives d ON a.drive_id = d.id 
           WHERE a.id = %s""",
        (app_id,),
        fetch_one=True
    )
    
    if not application:
        flash('Application not found.', 'error')
        return redirect(url_for('tpo_dashboard'))
    
    # Update status
    db.execute_query(
        "UPDATE applications SET status = %s WHERE id = %s",
        (status, app_id)
    )
    
    # Send email notification
    send_application_update_email(
        student_email=application['email'],
        student_name=application['student_name'],
        company_name=application['company_name'],
        job_role=application['job_role'],
        status=status
    )
    
    # Create notification
    db.execute_query(
        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
        (application['student_id'], f"Your application status updated to {status} for {application['company_name']}", 'info')
    )
    
    flash('Application status updated and email sent.', 'success')
    return redirect(url_for('tpo_dashboard'))

@app.route('/tpo/upload_offer_letter/<int:app_id>', methods=['POST'])
@login_required
@role_required('tpo')
def upload_offer_letter(app_id):
    """Upload offer letter for an application"""
    if 'offer_letter' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('tpo_dashboard'))
    
    file = request.files['offer_letter']
    if file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('tpo_dashboard'))
    
    if not allowed_file(file.filename):
        flash('Invalid file type. Please upload PDF or DOCX.', 'error')
        return redirect(url_for('tpo_dashboard'))
    
    # Get application
    application = db.execute_query(
        """SELECT a.*, u.email, u.name as student_name, d.company_name, d.job_role 
           FROM applications a 
           JOIN users u ON a.student_id = u.id 
           JOIN drives d ON a.drive_id = d.id 
           WHERE a.id = %s""",
        (app_id,),
        fetch_one=True
    )
    
    if not application:
        flash('Application not found.', 'error')
        return redirect(url_for('tpo_dashboard'))
    
    # Save file
    filename = secure_filename(f"offer_{app_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'offers', filename)
    file.save(file_path)
    
    # Update application status to Selected
    db.execute_query(
        "UPDATE applications SET status = 'Selected' WHERE id = %s",
        (app_id,)
    )
    
    # Save offer letter record
    db.execute_query(
        "INSERT INTO offer_letters (application_id, file_path, uploaded_by) VALUES (%s, %s, %s)",
        (app_id, file_path, session['user_id'])
    )
    
    # Send email with offer letter
    send_application_update_email(
        student_email=application['email'],
        student_name=application['student_name'],
        company_name=application['company_name'],
        job_role=application['job_role'],
        status='Selected',
        offer_letter_path=file_path
    )
    
    # Create notification
    db.execute_query(
        "INSERT INTO notifications (user_id, message, type) VALUES (%s, %s, %s)",
        (application['student_id'], f"Offer letter received from {application['company_name']}!", 'success')
    )
    
    flash('Offer letter uploaded and email sent successfully!', 'success')
    return redirect(url_for('tpo_dashboard'))

@app.route('/tpo/export_report')
@login_required
@role_required('tpo')
def export_tpo_report():
    """Export comprehensive placement report as Excel"""
    # Get all data
    students = db.execute_query(
        """SELECT u.*, 
                  COUNT(DISTINCT a.id) as total_applications,
                  COUNT(DISTINCT CASE WHEN a.status = 'Selected' THEN a.id END) as selected_count
           FROM users u
           LEFT JOIN applications a ON u.id = a.student_id
           WHERE u.role = 'student'
           GROUP BY u.id""",
        fetch_all=True
    )
    
    drives = db.execute_query(
        "SELECT * FROM drives ORDER BY created_at DESC",
        fetch_all=True
    )
    
    # Create Excel file with multiple sheets
    wb = openpyxl.Workbook()
    
    # Students sheet
    ws1 = wb.active
    ws1.title = "Students"
    ws1.append(['Name', 'Email', 'Department', 'Approved', 'Total Applications', 'Selected'])
    for student in students:
        ws1.append([
            student['name'],
            student['email'],
            student.get('department', ''),
            'Yes' if student['is_approved'] else 'No',
            student['total_applications'] or 0,
            student['selected_count'] or 0
        ])
    
    # Drives sheet
    ws2 = wb.create_sheet("Drives")
    ws2.append(['Company', 'Job Role', 'Eligibility', 'Last Date', 'Status'])
    for drive in drives:
        ws2.append([
            drive['company_name'],
            drive['job_role'],
            drive.get('eligibility', ''),
            str(drive['last_date']),
            drive['status']
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name='placement_report.xlsx')

# ==================== API Routes ====================

@app.route('/api/notifications/mark_read/<int:notif_id>', methods=['POST'])
@login_required
def mark_notification_read(notif_id):
    """Mark notification as read"""
    db.execute_query(
        "UPDATE notifications SET is_read = TRUE WHERE id = %s AND user_id = %s",
        (notif_id, session['user_id'])
    )
    return jsonify({'success': True})

# ==================== Error Handlers ====================

@app.errorhandler(404)
def not_found(error):
    return render_template('index.html', error='Page not found'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('index.html', error='Internal server error'), 500

# ==================== Main ====================

if __name__ == '__main__':
    # Initialize database on first run
    try:
        from init_db import init_database
        init_database()
    except Exception as e:
        print(f"Database initialization note: {e}")
    
    app.run(debug=True, host='0.0.0.0', port=5000)

